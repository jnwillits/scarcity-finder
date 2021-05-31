"""
Scarcity Finder
Copyright (c) 2021 (MIT License) Jeffrey Neil Willits   @jnwillits

This ulility scraps a big box's commercial site to get inventory quantities from selected products. It gets the
product URL's from and Excel file and files the inventory data to the same file. An algorithm will be added to 
provide alerts for items that are in short supply and may be good candidates for resale. It is work-in-progress...

"""

# import pandas as pd
import re
from bs4 import BeautifulSoup
from selenium import webdriver
import os
import sys
import openpyxl


def read_col(ws, start_row, start_col):
    col_list = []
    last_row = len(ws[start_col])
    for row in ws.iter_rows(min_row=start_row, min_col=start_col, max_row=last_row, max_col=start_col):
        for cell in row:
            if cell.value is not None:
                col_list.append(cell.value)
    return col_list


def read_row(ws, start_row, start_col):
    row_list = []
    last_col = len(ws[start_row])
    for col in ws.iter_cols(min_col=start_col, min_row=start_row, max_col=last_col, max_row=start_row):
        for cell in col:
            if cell.value is not None:
                row_list.append(cell.value)
    return row_list


if __name__ == '__main__':
    start_col = 1
    start_row = 1
    f_str = 'sf-data.xlsx'
    wb = openpyxl.load_workbook(f_str)
    ws = wb['Data']
    urls = read_col(ws, start_row, start_col)
    titles = []
    stores = []
    internet_nums = []
    inv_qty = []
    scrapping = []
    row = start_row
    data = ''
    for row in range(0, len(urls)):
        driver = webdriver.Chrome(executable_path='chromedriver.exe') 
        driver.get(urls[row])
        content = driver.page_source
        soup = BeautifulSoup(content, features="html.parser")
        
        # Get product name.
        title = soup.findAll(attrs={'class': 'product-details__title'})
        try:
            titles.append(re.findall(r'>(.+?)<', str(title))[0])
        except:
            titles.append('X')   

        # Get the internet number of the product for reference - last element in the list.
        try:
            internet_nums.append(re.findall('[0-9]+', urls[row])[-1])
        except:
            internet_nums.append 

        # Get on-hand inventory quantity.
        data = soup.findAll(attrs={'class': 'aislebay-wrapper--inventory'})

        if "Unavailable" in str(data):
            inv_qty.append('U')
        else:
            try:
                inv_qty.append(int(re.findall('\\b\\d+\\b', str(data))[0]))
            except:
                inv_qty.append('X') 

    # Read rows to a list, append scrapped data, and re-write the rows.
    for i in range(0, len(urls)):
        ws.cell(row=i+1, column=3).value = titles[i]
        ws.cell(row=i+1, column=4).value = internet_nums[i]
        row_data = read_row(ws, start_row + i, start_col)
        row_data.append(inv_qty[i])
        for c in range(0, len(row_data)): 
             ws.cell(row=i+1, column=1+c).value = row_data[c]
    wb.save(f_str)
    wb.close()
    driver.close()
sys.exit()
